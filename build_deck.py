# -*- coding: utf-8 -*-
# "Matching Marcas x Roster" — full deck, one section per brand, generated
# straight from the client's Excel (no hand-typed brand data, to avoid
# transcription errors across 21 brands).
import json, math, os, re, sys
import openpyxl

def brand_slug(marca):
    return re.sub(r'[^a-z0-9]+', '-', marca.lower())

# PDF export: each brand gets its own named @page sized to fit its own
# secondary-grid row count, so a brand with many matches (e.g. Samsung, 11)
# doesn't spill onto a mostly-blank continuation page.
PDF_PAGE_W = 1600
PDF_PAGE_BASE_H = 1140
PDF_PAGE_ROW_H = 256
PDF_GRID_COLS = 8

def brand_page_height(n_secondary):
    rows = math.ceil(n_secondary / PDF_GRID_COLS) if n_secondary else 1
    extra_rows = max(0, rows - 1)
    return PDF_PAGE_BASE_H + extra_rows * PDF_PAGE_ROW_H

HERE = os.path.dirname(__file__)
XLSX_PATH = os.path.join(HERE, "Matching marcas x roster.xlsx")

people = json.load(open(os.path.join(HERE, "people.json")))
fonts = json.load(open(os.path.join(HERE, "fonts.json")))

LOGO_SVG_RAW = open("/Users/tomascardozo/main/big/brand/nuevo/big-logo2.svg").read()
ISOTIPO_PATH = re.search(r'\sd="([^"]+)"', LOGO_SVG_RAW).group(1)
ISOTIPO_VIEWBOX = re.search(r'viewBox="([^"]+)"', LOGO_SVG_RAW).group(1)

FONDO_SVG_RAW = open(os.path.join(HERE, "assets", "vectorfondo.svg")).read()
FONDO_PATH = re.search(r'\sd="([^"]+)"', FONDO_SVG_RAW).group(1)
FONDO_VIEWBOX = re.search(r'viewBox="([^"]+)"', FONDO_SVG_RAW).group(1)

def fondo_svg():
    # same curve used on the roster site, recolored for a dark (azul) card
    # instead of the original black-on-peach treatment
    return (f'<svg viewBox="{FONDO_VIEWBOX}" preserveAspectRatio="xMidYMid slice">'
            f'<path d="{FONDO_PATH}" fill="#FFFFFF" fill-opacity="0.08"/></svg>')

IG_ICON = '''<svg width="15" height="15" viewBox="4.5 4.5 11 11" fill="none" xmlns="http://www.w3.org/2000/svg">
<path fill-rule="evenodd" clip-rule="evenodd" d="M10 13C11.6568 13 13 11.6568 13 10C13 8.34315 11.6568 7 10 7C8.34315 7 7 8.34315 7 10C7 11.6568 8.34315 13 10 13ZM10 12C11.1046 12 12 11.1046 12 10C12 8.89543 11.1046 8 10 8C8.89543 8 8 8.89543 8 10C8 11.1046 8.89543 12 10 12Z" fill="currentColor"/>
<path d="M13 6.5C12.7239 6.5 12.5 6.72386 12.5 7C12.5 7.27614 12.7239 7.5 13 7.5C13.2762 7.5 13.5 7.27614 13.5 7C13.5 6.72386 13.2762 6.5 13 6.5Z" fill="currentColor"/>
<path fill-rule="evenodd" clip-rule="evenodd" d="M4.82698 6.13803C4.5 6.77977 4.5 7.61985 4.5 9.3V10.7C4.5 12.3801 4.5 13.2202 4.82698 13.8619C5.1146 14.4264 5.57354 14.8854 6.13803 15.173C6.77977 15.5 7.61985 15.5 9.3 15.5H10.7C12.3801 15.5 13.2202 15.5 13.8619 15.173C14.4264 14.8854 14.8854 14.4264 15.173 13.8619C15.5 13.2202 15.5 12.3801 15.5 10.7V9.3C15.5 7.61985 15.5 6.77977 15.173 6.13803C14.8854 5.57354 14.4264 5.1146 13.8619 4.82698C13.2202 4.5 12.3801 4.5 10.7 4.5H9.3C7.61985 4.5 6.77977 4.5 6.13803 4.82698C5.57354 5.1146 5.1146 5.57354 4.82698 6.13803ZM10.7 5.5H9.3C8.44342 5.5 7.86113 5.50078 7.41104 5.53755C6.97262 5.57337 6.74842 5.6383 6.59202 5.71799C6.2157 5.90974 5.90974 6.2157 5.71799 6.59202C5.6383 6.74842 5.57337 6.97262 5.53755 7.41104C5.50078 7.86113 5.5 8.44342 5.5 9.3V10.7C5.5 11.5566 5.50078 12.1388 5.53755 12.5889C5.57337 13.0274 5.6383 13.2516 5.71799 13.408C5.90974 13.7843 6.2157 14.0902 6.59202 14.282C6.74842 14.3617 6.97262 14.4267 7.41104 14.4625C7.86113 14.4992 8.44342 14.5 9.3 14.5H10.7C11.5566 14.5 12.1388 14.4992 12.5889 14.4625C13.0274 14.4267 13.2516 14.3617 13.408 14.282C13.7843 14.0902 14.0902 13.7843 14.282 13.408C14.3617 13.2516 14.4267 13.0274 14.4625 12.5889C14.4992 12.1388 14.5 11.5566 14.5 10.7V9.3C14.5 8.44342 14.4992 7.86113 14.4625 7.41104C14.4267 6.97262 14.3617 6.74842 14.282 6.59202C14.0902 6.2157 13.7843 5.90974 13.408 5.71799C13.2516 5.6383 13.0274 5.57337 12.5889 5.53755C12.1388 5.50078 11.5566 5.5 10.7 5.5Z" fill="currentColor"/>
</svg>'''

TT_ICON = '''<svg width="15" height="15" viewBox="4.5 4.5 11 11" fill="none" xmlns="http://www.w3.org/2000/svg">
<path d="M12.9314 6.22392C12.5107 5.76167 12.2578 5.15893 12.2578 4.5H11.7306C11.8664 5.22023 12.3137 5.83829 12.9314 6.22392Z" fill="currentColor"/>
<path d="M8.05382 9.94764C7.14061 9.94764 6.3978 10.6602 6.3978 11.5362C6.3978 12.1467 6.75989 12.6778 7.28704 12.9434C7.09003 12.683 6.97288 12.3637 6.97288 12.0164C6.97288 11.1404 7.71569 10.4278 8.62891 10.4278C8.7993 10.4278 8.96435 10.4558 9.11879 10.5018V8.57106C8.95902 8.5506 8.79663 8.53782 8.62891 8.53782C8.59961 8.53782 8.57299 8.54038 8.5437 8.54038V10.0217C8.38662 9.97569 8.22421 9.94764 8.05382 9.94764Z" fill="currentColor"/>
<path d="M14.4249 7.07184V8.54038C13.4026 8.54038 12.4548 8.22626 11.6826 7.69501V11.5387C11.6826 13.4568 10.0559 15.0199 8.05377 15.0199C7.28168 15.0199 6.56283 14.7849 5.97444 14.389C6.63737 15.0709 7.58253 15.5 8.62887 15.5C10.6283 15.5 12.2577 13.9395 12.2577 12.0189V8.17516C13.0298 8.70641 13.9776 9.02051 15 9.02051V7.13061C14.8003 7.13061 14.6086 7.11016 14.4249 7.07184Z" fill="currentColor"/>
<path d="M11.6826 11.5387V7.69501C12.4548 8.22626 13.4026 8.54038 14.4249 8.54038V7.07184C13.8339 6.95183 13.3148 6.64535 12.9314 6.22392C12.3137 5.83829 11.869 5.22023 11.7279 4.5H10.2849L10.2822 12.0776C10.2503 12.9256 9.52076 13.6075 8.62887 13.6075C8.07507 13.6075 7.58785 13.3444 7.28434 12.946C6.75718 12.6778 6.3951 12.1492 6.3951 11.5387C6.3951 10.6628 7.13791 9.95017 8.05111 9.95017C8.2215 9.95017 8.38657 9.97828 8.541 10.0243V8.54293C6.58147 8.5838 5 10.1264 5 12.0189C5 12.9332 5.37008 13.7658 5.97444 14.389C6.56283 14.7849 7.28168 15.0199 8.05377 15.0199C10.0532 15.0199 11.6826 13.4568 11.6826 11.5387Z" fill="currentColor"/>
</svg>'''

def isotipo_svg(size=24, color="#FFFFFF"):
    return f'<svg width="{size}" height="{size}" viewBox="{ISOTIPO_VIEWBOX}" style="color:{color}"><path fill="currentColor" d="{ISOTIPO_PATH}"/></svg>'

# name (as written in the Excel) -> pid in people.json
NAME_TO_PID = {
    "Pelao Khe": "pelao_khe", "Juli Savioli": "juli_savioli", "Pia Scarnato": "pia_scarnato",
    "Benja Calero": "benja_calero", "Cris Pierri": "cris_pierri", "Dulce Pink": "dulce_pink",
    "Hablemos de Cine": "hablemos_de_cine", "Ber Scarnato": "ber_scarnato",
    "Renata Blasevich": "renata_blasevich", "Giuli Bellicoso": "giuli_bellicoso",
    "Mariano Bondar": "mariano_bondar", "Inachomer": "inachomer", "Santi Gallo": "santi_gallo",
    "Tiago Bergallo": "tiago_bergallo", "Maca Castro": "maca_castro", "Pauli Veltrano": "pauli_veltrano",
    "Agus FC": "agustina_cambra", "Eve Vidal": "eve_vidal", "Pablo Bruschi": "pablo_bruschi",
    "Inez Mumy": "inez", "Mumy Ratibel": "mumy", "Joselo Márquez": "joselo_marquez",
    "Bruno Rondini": "bruno_rondini", "LuBru Invierte": "lubru_invierte",
    "Los Arias Brothers": "los_arias_brothers", "Ammichis": "ammichis", "Facu Garcia": "facu_garcia",
    "Giuli Lourdes": "giuli_lourdes", "Martu Morales": "martu_morales", "Tomas Alvarez": "tomas_alvarez",
    "Tomás Alvarez": "tomas_alvarez", "Fran Silva": "fran_silva", "El Capo Willy": "el_capo_willy",
    "Agus Benca": "agus_benca", "Hilario José": "hilario_jose", "Yo Soy Brisa": "yo_soy_brisa",
    "Nanu Yael": "nanu_yael", "Sabri Ludmila": "sabri_ludmila", "Soy Dalto": "soy_dalto",
    "Lucas Monopoli": "lucas_monopoli", "Nico Grasso": "nico_grasso", "Gena Pedrazzoli": "gena_pedrazzoli",
    "Gena Pedrazzi": "gena_pedrazzoli", "Mely Francano": "mely_francano", "Lean Riccio": "lean_riccio",
}

# hero justificación that reads oddly outside the influencer's own primary
# category — trimmed to the generic part, per user decision (2026-08-20)
JUSTIFICACION_OVERRIDE = {
    ("Cif", "eve_vidal"): "Cocina y bienestar.",
    ("Molinos Ala", "eve_vidal"): "Cocina y bienestar.",
    ("Secret", "giuli_bellicoso"): "Vlogs, outfits, baile y belleza.",
}

# ---- load the Excel ----
wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
ws1 = wb["Influencer → Marcas"]
ws2 = wb["Marca → Influencers"]

influencer_info = {}
for row in ws1.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    influencer_info[row[0].strip()] = dict(
        ig_handle=(row[1] or "").strip(), match_principal=(row[2] or "").strip(),
        otras_marcas=(row[3] or "").strip(), justificacion=(row[4] or "").strip(),
    )

BRANDS = []
for row in ws2.iter_rows(min_row=2, values_only=True):
    if not row[0]:
        continue
    marca, cliente, categoria, influencers_raw = row
    names = [n.strip() for n in influencers_raw.split(";") if n.strip()]
    BRANDS.append(dict(marca=marca.strip(), cliente=(cliente or "").strip(),
                        categoria=(categoria or "").strip(), names=names))

def secondary_card(pid):
    p = people[pid]
    ig_href = p["ig_url"] or "#"
    tt_href = p.get("tt_url") or "#"
    tt_chip = (
        f'<a class="mcard-sec-chip" href="{tt_href}" target="_blank" rel="noopener">{TT_ICON}<span>{p["tt"]}</span></a>'
        if p.get("tt") else ""
    )
    return f'''
    <div class="mcard-sec">
      <img class="mcard-sec-photo" src="{p["photo"]}" alt="{p["name"]}" loading="lazy" decoding="async" width="200" height="240">
      <div class="mcard-sec-body">
        <h4 class="mcard-sec-name">{p["name"]}</h4>
        <div class="mcard-sec-chips">
          <a class="mcard-sec-chip" href="{ig_href}" target="_blank" rel="noopener">{IG_ICON}<span>{p["ig"] or "—"}</span></a>
          {tt_chip}
        </div>
      </div>
    </div>'''

def brand_section(brand):
    hero_name = brand["names"][0]
    hero_pid = NAME_TO_PID[hero_name]
    hero = people[hero_pid]
    info = influencer_info.get(hero_name, {})
    justificacion = JUSTIFICACION_OVERRIDE.get(
        (brand["marca"], hero_pid), info.get("justificacion", "")
    )
    secondary_pids = [NAME_TO_PID[n] for n in brand["names"][1:]]
    secondary_html = "".join(secondary_card(pid) for pid in secondary_pids)

    hero_href = hero["ig_url"] or "#"
    hero_tt_href = hero.get("tt_url") or "#"
    hero_tt_chip = (
        f'<a class="hero-photo-chip" href="{hero_tt_href}" target="_blank" rel="noopener">{TT_ICON}<span>{hero["tt"]}</span></a>'
        if hero.get("tt") else ""
    )
    slug = brand_slug(brand["marca"])
    return f'''
<section class="brand-section" id="brand-{slug}" style="page: page-{slug};">
  <div class="brand-badge-row">
    {isotipo_svg(18, "#FFFFFF")}
    <span class="brand-badge-x">×</span>
    <img class="brand-badge-logo" src="assets/logomedialab.avif" alt="Medialab">
  </div>
  <div class="brand-header">
    <div>
      <div class="brand-eyebrow">{brand["cliente"]}</div>
      <h1 class="brand-title">{brand["marca"]}</h1>
    </div>
    <span class="brand-category">{brand["categoria"]}</span>
  </div>

  <div class="hero-row">
    <div class="hero-photo-card">
      <img src="{hero["photo"]}" alt="{hero["name"]}" loading="lazy" decoding="async">
      <div class="hero-photo-scrim" aria-hidden="true"></div>
      <div class="hero-photo-top-scrim" aria-hidden="true"></div>
      <span class="hero-photo-match">★ Match principal</span>
      <div class="hero-photo-tag">
        <div class="hero-photo-name">{hero["name"]}</div>
        <div class="hero-photo-chips">
          <a class="hero-photo-chip" href="{hero_href}" target="_blank" rel="noopener">{IG_ICON}<span>{hero["ig"] or "—"}</span></a>
          {hero_tt_chip}
        </div>
      </div>
    </div>
    <div class="hero-text-card">
      <div class="hero-text-bg" aria-hidden="true">{fondo_svg()}</div>
      <div class="hero-text-eyebrow">Por qué la recomendamos</div>
      <p class="hero-text-quote">{justificacion}</p>
    </div>
  </div>

  <div class="secondary-head">
    <div class="secondary-eyebrow">También recomendados para {brand["marca"]}</div>
  </div>
  <div class="secondary-grid">{secondary_html}</div>
</section>'''

font_faces = "\n".join(f'''
@font-face {{
  font-family: 'Inter';
  font-weight: {w};
  font-style: normal;
  src: url(data:font/woff2;base64,{fonts[name]}) format('woff2');
  font-display: swap;
}}''' for name, w in [("Regular", 400), ("Medium", 500), ("Bold", 700), ("Black", 900)])

nav_links = "\n".join(
    f'<a href="#brand-{brand_slug(b["marca"])}" class="nav-link">{b["marca"]}</a>'
    for b in BRANDS
)

index_items = "\n".join(f'''
    <a class="index-item" href="#brand-{brand_slug(b["marca"])}">
      <span class="index-item-num">{i+1:02d}</span>
      <span class="index-item-body">
        <span class="index-item-name">{b["marca"]}</span>
        <span class="index-item-cat">{b["categoria"]}</span>
      </span>
    </a>''' for i, b in enumerate(BRANDS))

index_section = f'''
<section class="index-section">
  <div class="index-eyebrow">Índice</div>
  <h1 class="index-title">Marcas</h1>
  <div class="index-grid">{index_items}</div>
</section>'''

sections_html = "\n".join(brand_section(b) for b in BRANDS)

# PDF export: named @page rule per brand, sized to that brand's own
# secondary-grid row count (see brand_page_height above).
brand_page_rules = "\n".join(
    f'@page page-{brand_slug(b["marca"])} {{ size: {PDF_PAGE_W}px {brand_page_height(len(b["names"]) - 1)}px; margin: 0; }}'
    for b in BRANDS
)

html = f'''<title>Matching Marcas x Roster — BIG Agency</title>
<style>
{font_faces}

:root {{
  --azul: #33419A;
  --naranja: #F36F2C;
  --fondo: #F7D8BD;
  --lima: #E8F29C;
  --negro: #0D0D14;
  --blanco: #FFFFFF;
}}
* {{ box-sizing: border-box; margin: 0; padding: 0; }}
html {{ scroll-behavior: smooth; }}
body {{
  font-family: 'Inter', -apple-system, sans-serif;
  background: var(--fondo); color: var(--azul);
  letter-spacing: -0.01em;
}}

/* ---------- nav (screen only, hidden on print) ---------- */
.deck-nav {{
  position: sticky; top: 0; z-index: 50; background: var(--fondo);
  border-bottom: 1px solid rgba(51,65,154,.15);
  display: flex; align-items: center; gap: 10px; padding: 14px 24px;
  overflow-x: auto; white-space: nowrap;
}}
.nav-link {{
  background: var(--fondo); color: var(--naranja); text-decoration: none; font-weight: 600; font-size: 12.5px;
  letter-spacing: .02em; text-transform: uppercase; padding: 7px 13px;
  border: 1.5px solid var(--naranja); border-radius: 999px; flex-shrink: 0;
  transition: background .2s ease, color .2s ease, border-color .2s ease, transform .2s ease;
}}
.nav-link:hover {{ background: var(--naranja); color: var(--blanco); border-color: transparent; transform: translateY(-2px); }}

/* ---------- brand section ---------- */
.brand-section {{
  min-height: 100vh; padding: 56px 64px 72px; position: relative; overflow: hidden;
}}
.brand-section:nth-of-type(even) {{ background: #F2CBA7; }}
.brand-badge-row {{
  display: inline-flex; align-items: center; gap: 14px;
  background: var(--naranja); border-radius: 14px; padding: 9px 18px 9px 9px; margin-bottom: 32px;
}}
.brand-badge-x {{ color: var(--blanco); font-weight: 900; font-size: 15px; opacity: .75; }}
.brand-badge-logo {{ height: 18px; width: auto; display: block; }}
.brand-header {{ display: flex; align-items: flex-end; justify-content: space-between; gap: 24px; margin-bottom: 40px; flex-wrap: wrap; }}
.brand-eyebrow {{ color: var(--naranja); font-weight: 700; font-size: 13px; letter-spacing: .08em; text-transform: uppercase; }}
.brand-title {{
  color: var(--azul); font-weight: 900; text-transform: uppercase;
  font-size: clamp(48px, 6vw, 88px); line-height: .88; letter-spacing: -0.05em; margin-top: 6px;
}}
.brand-category {{
  display: inline-flex; align-items: center; padding: 10px 18px; border-radius: 999px;
  background: var(--naranja); color: var(--blanco); font-weight: 700; font-size: 13px;
  letter-spacing: .02em; text-transform: uppercase; white-space: nowrap;
}}

.hero-row {{ display: flex; gap: 32px; align-items: stretch; flex-wrap: wrap; }}
.hero-photo-card {{
  flex: 0 0 340px; height: 460px; position: relative; border-radius: 32px; overflow: hidden; background: #d9d9d9;
  display: block; text-decoration: none;
  transition: transform .4s cubic-bezier(.16,1,.3,1), box-shadow .4s cubic-bezier(.16,1,.3,1);
}}
.hero-photo-card:hover {{ transform: translateY(-6px); box-shadow: 0 22px 44px rgba(13,13,23,.28); }}
.hero-photo-card img {{
  position: absolute; inset: 0; width: 100%; height: 100%; object-fit: cover; object-position: center top;
  transition: transform .5s cubic-bezier(.16,1,.3,1);
}}
.hero-photo-card:hover img {{ transform: scale(1.05); }}
.hero-photo-scrim {{
  position: absolute; inset: 0; pointer-events: none;
  background: linear-gradient(to top, rgba(13,13,23,.75) 0%, rgba(13,13,23,0) 40%);
}}
.hero-photo-top-scrim {{
  position: absolute; top: 0; left: 0; right: 0; height: 120px; pointer-events: none;
  background: linear-gradient(to bottom, rgba(13,13,23,.55) 0%, rgba(13,13,23,0) 100%);
}}
.hero-photo-tag {{ position: absolute; left: 24px; bottom: 24px; right: 24px; color: var(--blanco); }}
.hero-photo-match {{
  position: absolute; top: 24px; left: 24px; z-index: 1;
  display: inline-flex; align-items: center; gap: 6px; width: fit-content;
  padding: 6px 12px; border-radius: 999px; background: var(--naranja); color: var(--blanco);
  font-weight: 700; font-size: 11px; text-transform: uppercase; letter-spacing: .03em;
}}
.hero-photo-name {{ font-weight: 700; font-size: 26px; letter-spacing: -0.03em; }}
.hero-photo-chips {{ display: flex; flex-wrap: wrap; gap: 8px; margin-top: 10px; }}
.hero-photo-chip {{
  display: inline-flex; align-items: center; gap: 6px;
  padding: 7px 12px; border-radius: 999px; background: rgba(255,255,255,.22);
  -webkit-backdrop-filter: blur(4px); backdrop-filter: blur(4px);
  color: var(--blanco); text-decoration: none;
  font-weight: 700; font-size: 13px;
  transition: background .2s ease, transform .2s ease;
}}
.hero-photo-chip:hover {{ background: var(--naranja); transform: translateY(-2px); }}

.hero-text-card {{
  flex: 1; min-width: 320px; background: var(--azul); border-radius: 32px; padding: 44px;
  display: flex; flex-direction: column; justify-content: center; gap: 18px;
  color: var(--blanco); position: relative; overflow: hidden;
}}
.hero-text-bg {{ position: absolute; inset: 0; pointer-events: none; overflow: hidden; }}
.hero-text-bg svg {{ position: absolute; top: -10%; left: -10%; width: 120%; height: 120%; display: block; }}
.hero-text-eyebrow {{ position: relative; z-index: 1; color: var(--lima); font-weight: 700; font-size: 13px; letter-spacing: .08em; text-transform: uppercase; }}
.hero-text-quote {{ position: relative; z-index: 1; font-weight: 700; font-size: 30px; line-height: 1.28; letter-spacing: -0.02em; max-width: 640px; }}

.secondary-head {{ margin: 56px 0 24px; }}
.secondary-eyebrow {{ color: var(--naranja); font-weight: 700; font-size: 13px; letter-spacing: .08em; text-transform: uppercase; }}
.secondary-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(160px, 1fr)); gap: 16px; }}
.mcard-sec {{
  display: block; position: relative; border-radius: 22px; overflow: hidden; background: #d9d9d9; height: 240px;
  transition: transform .35s cubic-bezier(.16,1,.3,1), box-shadow .35s cubic-bezier(.16,1,.3,1);
}}
.mcard-sec:hover {{ transform: translateY(-4px); box-shadow: 0 16px 30px rgba(13,13,23,.25); }}
.mcard-sec-photo {{
  position: absolute; inset: 0; width: 100%; height: 100%; object-fit: cover; object-position: center top;
  transition: transform .45s cubic-bezier(.16,1,.3,1);
}}
.mcard-sec:hover .mcard-sec-photo {{ transform: scale(1.06); }}
.mcard-sec-body {{
  position: absolute; left: 0; right: 0; bottom: 0; padding: 14px;
  background: linear-gradient(to top, rgba(13,13,23,.8), rgba(13,13,23,0));
  display: flex; flex-direction: column; gap: 6px;
}}
.mcard-sec-name {{ color: var(--blanco); font-weight: 700; font-size: 14px; letter-spacing: -0.02em; }}
.mcard-sec-chips {{ display: flex; flex-wrap: wrap; gap: 5px; }}
.mcard-sec-chip {{
  display: inline-flex; align-items: center; gap: 3px; text-decoration: none;
  padding: 4px 7px; border-radius: 999px; background: rgba(255,255,255,.22);
  -webkit-backdrop-filter: blur(3px); backdrop-filter: blur(3px);
  color: var(--blanco); font-size: 10.5px; font-weight: 700;
  transition: background .2s ease, transform .2s ease;
}}
.mcard-sec-chip:hover {{ background: var(--naranja); transform: translateY(-2px); }}
.mcard-sec-chip svg {{ display: block; width: 11px; height: 11px; }}

.cover-section {{
  min-height: 100vh; position: relative; overflow: hidden; background: var(--azul);
  display: flex; flex-direction: column; align-items: center; justify-content: center; gap: 32px;
  text-align: left; padding: 64px 96px;
}}
.cover-bg {{ position: absolute; inset: 0; pointer-events: none; overflow: hidden; }}
.cover-bg svg {{ position: absolute; top: -10%; left: -10%; width: 120%; height: 120%; display: block; }}
.cover-lockup {{
  position: relative; z-index: 1; width: min(760px, 78vw); height: auto; display: block;
}}

/* ---------- índice (PDF export only) ---------- */
.index-section {{
  min-height: 100vh; padding: 72px 96px; position: relative;
  background: var(--fondo);
  display: flex; flex-direction: column; justify-content: center; gap: 40px;
}}
.index-eyebrow {{ color: var(--naranja); font-weight: 700; font-size: 14px; letter-spacing: .08em; text-transform: uppercase; }}
.index-title {{
  color: var(--azul); font-weight: 900; text-transform: uppercase;
  font-size: clamp(40px, 5vw, 72px); line-height: .88; letter-spacing: -0.05em; margin-top: -24px;
}}
.index-grid {{ display: grid; grid-template-columns: repeat(auto-fill, minmax(280px, 1fr)); gap: 14px; }}
.index-item {{
  display: flex; align-items: baseline; gap: 14px; text-decoration: none; color: var(--azul);
  background: var(--blanco); border-radius: 14px; padding: 16px 18px;
  border: 1.5px solid rgba(51,65,154,.12);
  transition: transform .2s ease, border-color .2s ease;
}}
.index-item:hover {{ transform: translateY(-3px); border-color: var(--naranja); }}
.index-item-num {{ font-weight: 700; font-size: 13px; color: var(--naranja); flex-shrink: 0; }}
.index-item-body {{ display: flex; flex-direction: column; gap: 3px; }}
.index-item-name {{ font-weight: 800; font-size: 16.5px; letter-spacing: -0.02em; }}
.index-item-cat {{ font-weight: 600; font-size: 11px; color: rgba(51,65,154,.6); text-transform: uppercase; letter-spacing: .03em; }}

@page {{
  size: {PDF_PAGE_W}px {PDF_PAGE_BASE_H}px;
  margin: 0;
}}
{brand_page_rules}

@media print {{
  * {{ -webkit-print-color-adjust: exact !important; print-color-adjust: exact !important; }}
  .deck-nav {{ display: none; }}
  .cover-section {{ min-height: 100vh; page-break-after: always; }}
  .index-section {{ min-height: 100vh; page-break-after: always; }}
  .brand-section {{
    min-height: 100vh; padding: 48px 64px 40px; page-break-after: always;
  }}
  .brand-badge-row {{ margin-bottom: 22px; }}
  .brand-header {{ margin-bottom: 28px; }}
  .secondary-head {{ margin: 36px 0 20px; }}
  .hero-photo-card:hover, .mcard-sec:hover, .index-item:hover,
  .hero-photo-chip:hover, .mcard-sec-chip:hover {{ transform: none; }}
}}
</style>

<section class="cover-section">
  <div class="cover-bg" aria-hidden="true">{fondo_svg()}</div>
  <img class="cover-lockup" src="assets/decktitle.svg" alt="BIG × Media Lab — Matching de marcas con creadores">
</section>

{{nav_or_index}}
{sections_html}
'''

for_pdf = "--pdf" in sys.argv

if not for_pdf:
    out_html = html.replace("{nav_or_index}", f'<nav class="deck-nav">{nav_links}</nav>')
    out_path = os.path.join(HERE, "index.html")
    with open(out_path, "w") as f:
        f.write(out_html)
    print("wrote", out_path, len(out_html) / 1024, "KB —", len(BRANDS), "brands")
else:
    out_html = html.replace("{nav_or_index}", index_section)
    out_path = os.path.join(HERE, "_print_source.html")
    with open(out_path, "w") as f:
        f.write(out_html)
    print("wrote", out_path, len(out_html) / 1024, "KB —", len(BRANDS), "brands (PDF source)")
